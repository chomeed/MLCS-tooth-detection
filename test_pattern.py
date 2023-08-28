from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Create a new workbook and select the active worksheet
workbook = Workbook()
sheet = workbook.active

# Define a pattern fill for red and green
red_fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
green_fill = PatternFill(start_color='00008000', end_color='00008000', fill_type='solid')

# Apply pattern fill to cells
sheet.cell(row=1, column=1).value = "Red Fill"
sheet.cell(row=1, column=1).fill = red_fill

sheet.cell(row=2, column=1).value = "Green Fill"
sheet.cell(row=2, column=1).fill = green_fill

# Save the workbook
workbook.save("pattern_fill_test.xlsx")
