from openpyxl import Workbook
from openpyxl.styles import PatternFill
import json

def compare_values_and_get_pattern_fill(value1, value2):
    if value1 == 'nan' or value2 == 'nan':
        return PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')

    red_fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
    green_fill = PatternFill(start_color='00008000', end_color='00008000', fill_type='solid')
    
    if value1 > value2:
        return green_fill
    elif value1 < value2:
        return red_fill
    else: 
        return PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')


# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
# sheet = write_wb.create_sheet('no name')

# Sheet1에다 입력
sheet = write_wb.active

# 데이터 불러오기
headers = ['category', 'mAP', 'mAP_50', 'mAP_75', 'mAP_s', 'mAP_m', 'mAP_l']
sheet.append(headers) 

# Baseline 불러오기 
with open('baseline.json', 'r') as rf: 
    baseline_data = json.load(rf)
    # data = list(data.items())

with open('test_json.json', 'r') as rf:
    data = json.load(rf)
    data = list(data.items())
    sorted_data = sorted(data, key=lambda x: x[0][-2:])
    print(sorted_data)

for row_idx, d in enumerate(sorted_data, start=2):
    cat = d[0]
    mAP_data = list(d[1].values())

    patterns = [PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')]

    for cur, base in zip(mAP_data, baseline_data[cat].values()):
        pattern_fill = compare_values_and_get_pattern_fill(cur, base)
        patterns.append(pattern_fill)
    
    new_data = mAP_data
    new_data.insert(0, cat)
    # new data is a list 

    alphabet = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    for col_idx, elem, pat in zip(alphabet, new_data, patterns):
        sheet[f"{col_idx}{row_idx}"] = elem
        sheet[f"{col_idx}{row_idx}"].fill = pat

    # sheet.append(new_data) 

# # #셀 단위로 추가
# # sheet.cell(5, 5, '5행5열')

write_wb.save("test3.xlsx")